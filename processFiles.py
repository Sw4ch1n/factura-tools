import os
import re
import shutil
import subprocess
import getpass
import xml.etree.ElementTree as ET
import pandas as pd
from unidecode import unidecode
import flet as ft
import logging
import openpyxl
from openpyxl.utils import get_column_letter

# Configuración del logger
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Obtener usuario actual
usuario = getpass.getuser()

def replace_chars(text):
    return unidecode(text) if text else text

def search_data(xml_file):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()

        fecha_timbrado = None
        emisor_nombre = None
        uuid = None
        total = None
        uso_cfdi = None
        
        for elem in root.iter():
            if elem.tag.endswith('TimbreFiscalDigital') and 'FechaTimbrado' in elem.attrib:
                fecha_timbrado = elem.attrib['FechaTimbrado']
            elif elem.tag.endswith('AlternateReference') and 'Date' in elem.attrib:
                fecha_timbrado = elem.attrib['Date']   

            if elem.tag.endswith('Emisor') and 'Nombre' in elem.attrib:
                emisor_nombre = elem.attrib['Nombre']
            elif elem.tag.endswith('Supplier') and 'Company' in elem.attrib:
                emisor_nombre = elem.attrib['Company']
            elif elem.tag.endswith('Emisor') and 'nombre' in elem.attrib:
                emisor_nombre = elem.attrib['nombre']
            elif elem.tag.endswith('Sender') and 'Company' in elem.attrib:
                emisor_nombre = elem.attrib['Company']
            elif elem.tag.endswith('Sender') and 'Role' in elem.attrib and elem.attrib['Role'] == 'supplier':
                emisor_nombre = elem.attrib['Name']
            elif elem.tag.endswith('Supplier') and 'Name' in elem.attrib:
                emisor_nombre = elem.attrib['Name']
            elif elem.tag.endswith('Issuer') and 'Name' in elem.attrib:
                emisor_nombre = elem.attrib['Name']
            
            if elem.tag.endswith('TimbreFiscalDigital') and 'UUID' in elem.attrib:
                uuid = elem.attrib['UUID']
            elif elem.tag.endswith('AlternateReference') and 'Ref' in elem.attrib:
                uuid = elem.attrib['Ref']
            elif elem.tag.endswith('Complemento'):
                for sub_elem in elem:
                    if sub_elem.tag.endswith('TimbreFiscalDigital') and 'UUID' in sub_elem.attrib:
                        uuid = sub_elem.attrib['UUID']
                        break

            if 'Total' in elem.attrib:
                total = elem.attrib['Total']
            elif 'total' in elem.attrib:
                total = elem.attrib['total']
            elif elem.tag.endswith('Comprobante') and 'Total' in elem.attrib:
                total = elem.attrib['Total']
            elif elem.tag.endswith('Invoice') and 'Total' in elem.attrib:
                total = elem.attrib['Total']

            if 'UsoCFDI' in elem.attrib:
                uso_cfdi = elem.attrib['UsoCFDI']
            elif 'Uso' in elem.attrib:
                uso_cfdi = elem.attrib['Uso']
            elif 'Use' in elem.attrib:
                uso_cfdi = elem.attrib['Use']
        
        return fecha_timbrado, emisor_nombre, total, uso_cfdi, uuid

    except ET.ParseError as e:
        logging.error(f"Error al analizar XML {xml_file}: {e}")
        return None, None, None, None, None
    except Exception as e:
        logging.error(f"Error desconocido al procesar {xml_file}: {e}")
        return None, None, None, None, None


def format_filename(fecha_timbrado, emisor_nombre, uuid):
    fecha = fecha_timbrado.split('T')[0] if fecha_timbrado else "00-00-0000"
    anio, mes, dia = fecha.split('-') if '-' in fecha else ("00", "00", "0000")
    uuid_f = (uuid[-6:].upper() if uuid else "NoEncontrado")

    emisor_nombre = replace_chars(emisor_nombre) if emisor_nombre else "NoEncontrado"
    emisor_file = re.sub(r"[ .,\-;&+'']", "", emisor_nombre)[:30].upper()

    return f"{dia}.{mes}.{anio}_{emisor_file}_{uuid_f}_FACTURA"

def rename_files(ruta_carpeta, nombre_archivo, nuevo_nombre, uuid):
    try:
        original_name = os.path.splitext(nombre_archivo)[0]
        original_route_xml = os.path.join(ruta_carpeta, f"{original_name}.xml")
        uuid_formated = uuid.replace("-", "").upper()
        ruta_uuid_pdf = os.path.join(ruta_carpeta, f"{uuid_formated}.pdf")

        new_name_xml = os.path.join(ruta_carpeta, f"{nuevo_nombre}.xml")
        new_name_pdf = os.path.join(ruta_carpeta, f"{nuevo_nombre}.pdf")
        suffix = 1
        while os.path.exists(new_name_xml) or os.path.exists(new_name_pdf):
            suffix += 1
            new_name_xml = os.path.join(ruta_carpeta, f"dp{suffix}_{nuevo_nombre}.xml")
            new_name_pdf = os.path.join(ruta_carpeta, f"dp{suffix}_{nuevo_nombre}.pdf")

        if os.path.exists(original_route_xml):
            shutil.move(original_route_xml, new_name_xml)
        if os.path.exists(ruta_uuid_pdf):
            shutil.move(ruta_uuid_pdf, new_name_pdf)

        return True
    except (FileNotFoundError, PermissionError) as e:
        logging.error(f"No se pudo renombrar: {str(e)}")
        return False

def processFiles(page: ft.Page, carpeta, sucursal, progress_bar: ft.ProgressBar):
    def mostrar_alerta(mensaje, titulo="Información"):
        dlg = ft.AlertDialog(
            title=ft.Text(titulo),
            content=ft.Text(mensaje),
            actions=[ft.TextButton("OK", on_click=lambda e: (setattr(dlg, "open", False), page.update()))]
        )
        page.open(dlg)
        page.update()

    try:
        archivos_xml = [f for f in os.listdir(carpeta) if f.lower().endswith('.xml')]
        total_archivos = len(archivos_xml)
        if not archivos_xml:
            mostrar_alerta("No se encontraron archivos XML en la carpeta.")
            return

        data = []

        for i, archivo in enumerate(archivos_xml):
            ruta_xml = os.path.join(carpeta, archivo)
            fecha_timbrado, emisor_nombre, total, uso_cfdi, uuid = search_data(ruta_xml)

            if not any([fecha_timbrado, emisor_nombre, total, uso_cfdi, uuid]):
                logging.warning(f"No se encontraron datos relevantes en el archivo XML: {archivo}")
                continue

            nuevo_nombre = format_filename(fecha_timbrado, emisor_nombre, uuid)
            if rename_files(carpeta, archivo, nuevo_nombre, uuid):
                logging.info(f"Archivos renombrados: {archivo} -> {nuevo_nombre}")
            else:
                logging.warning(f"No se pudieron renombrar los archivos para: {archivo}")

            if fecha_timbrado:
                fecha = fecha_timbrado.split('T')[0]
                anio, mes, dia = fecha.split('-')
                fecha_formateada = f"{dia}.{mes}.{anio}"
            else:
                fecha_formateada = "No encontrado"

            data.append({
                "FECHA": fecha_formateada,
                "NOMBRE DEL EMISOR (RAZON SOCIAL)": emisor_nombre if emisor_nombre else "No encontrado",
                "LOCALIDAD": sucursal,
                "IMPORTE": total if total else "No encontrado",
                "PROYECTO / EXPEDIENTE": "ND",
                "CONCEPTO GASTO": "ND",
                "UUID (FOLIO FISCAL)": uuid if uuid else "No encontrado",
                "USO DE CFDI": uso_cfdi if uso_cfdi else "No encontrado",
                "NOMBRE ORIGINAL XML": archivo,
                "NUEVO NOMBRE": nuevo_nombre,
            })

            porcentaje_avance = (i + 1) / total_archivos
            progress_bar.value = porcentaje_avance
            page.update()

        if data:
            ruta_excel = os.path.join(carpeta, "_REPORTE_DE_FACTURAS_RENOMBRADAS.xlsx")
            df = pd.DataFrame(data)
            df.to_excel(ruta_excel, index=False, engine='openpyxl')

            workbook = openpyxl.load_workbook(ruta_excel)
            sheet = workbook.active
            if sheet:
                for column in sheet.columns:
                    max_length = 0
                    if column and column[0] and column[0].column is not None:
                        column_name = get_column_letter(int(column[0].column))
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        sheet.column_dimensions[column_name].width = adjusted_width

                workbook.save(ruta_excel)
                mostrar_alerta("Renombrado de facturas completado. \n\nReporte en Excel generado en la carpeta.")
                try:
                    subprocess.Popen([ruta_excel], shell=True)
                except Exception as ex:
                    logging.error(f"No se pudo abrir el archivo Excel: {str(ex)}")
                    mostrar_alerta(f"No se pudo abrir el archivo Excel: {str(ex)}", "Error")

            else:
                logging.error("No se pudo acceder a la hoja de cálculo en el archivo Excel.")
                mostrar_alerta("Error al acceder al archivo Excel.", "Error")
            return ruta_excel
        else:
            mostrar_alerta("No se encontraron archivos XML para procesar.")
            return None

    except Exception as e:
        logging.error(f"Error al procesar archivos: {str(e)}")
        mostrar_alerta(f"Error al procesar archivos: {str(e)}", "Error")